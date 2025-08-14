import openpyxl
import os

# Thư mục chứa các file Excel
folder_path = 'E:\\CRUD\\input'

# Tạo một workbook mới để chứa dữ liệu gộp
merged_workbook = openpyxl.Workbook()

# Duyệt qua tất cả các tệp Excel trong thư mục
sheet_names = []

for root, dirs, files in os.walk(folder_path):
    for file_name in files:
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(root, file_name)

            # Đọc từng file Excel đầu vào
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet_names.append(sheet_name)

# Sắp xếp danh sách tên sheet theo thứ tự
sheet_names.sort()

# Tạo các sheet mới theo thứ tự đã sắp xếp
for sheet_name in sheet_names:
    new_sheet = merged_workbook.create_sheet(title=sheet_name)

# Xóa sheet mặc định được tạo ra khi tạo workbook mới
if 'Sheet' in merged_workbook.sheetnames:
    default_sheet = merged_workbook['Sheet']
    merged_workbook.remove(default_sheet)

# Lưu workbook mới chứa dữ liệu gộp và sheet đã sắp xếp
merged_workbook.save('merged_file.xlsx')
